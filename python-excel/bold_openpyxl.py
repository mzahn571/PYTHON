import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font

newfile = 'C:/Users/ab21600/Documents/DA/CFPB.xlsx'
wb = load_workbook(filename=newfile)
ws = wb['CFPB']
bold_font = Font(bold=True)

# Enumerate the cells in the second row
for cell in ws["1:1"]:
    cell.font = bold_font

wb.save(filename=newfile)
