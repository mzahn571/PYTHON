import openpyxl

wb = openpyxl.load_workbook('C:/Users/ab21600/Documents/Python-Remedy-Report/test1.xlsx')
#wb = openpyxl.load_workbook('C:/Users/ab21600/Documents/DA/CFPB.xlsx')
#ws = wb.get_sheet_by_name('CFPB')
ws = wb.get_sheet_by_name('test')
mylist = []
for row in ws.iter_rows('B{}:B{}'.format(ws.min_row,ws.max_row)):
    for cell in row:
        mylist.append(cell.value)
print(('\n').join(mylist)) 
