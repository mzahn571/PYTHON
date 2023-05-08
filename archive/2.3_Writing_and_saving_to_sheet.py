#!/usr/bin/env python3
"""
Creating a Workbook object with OpenPyXL
Adding, deleting and renaming sheets
Saving the workbook to disk 
"""
from openpyxl import Workbook

def main():
    #Creating the Workbook object 
    wb = Workbook()
    
    # Create s new sheets
    ws = wb.create_sheet("A sheet", 0) # insert at first position
    
    # Change the name of the sheet
    ws.title = "Hello World!"
    
    #Create another sheet 
    ws2 = wb.create_sheet("Sheet nr 2") # Insert the new sheet at the end
    
    #Create what sheets exists in the workbook
    print("Sheets in workbook:")
    for sheet in wb:
        print(sheet.title)
    print("-"*20)
    # Delete sheet "Sheet"
    wb.remove(wb["Sheet"])
    #del wb["Sheet"] also works
    
    # Check what sheets are left
    print("Sheets in workbook after deletion:")
    for sheet in wb:
        print(sheet.title)
        
    # Save the workbook to discard
    wb.save('2.3_Hello_sheets.xlsx')
    print("Existing main()")
    
if __name__ == " __main__":
    main()
    