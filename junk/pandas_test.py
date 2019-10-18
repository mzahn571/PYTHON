import openpyxl
import shutil
import copy
import pandas as pd
import os
import sys
import pathlib

import openpyxl as xl

path1 = 'C:/Users/ab21600/Documents/example1.xlsx'
path2 = 'C:/Users/ab21600/Documents/example2.xlsx'

wb1 = xl.load_workbook(filename=path1)
ws1 = wb1.worksheets[0]

wb2 = xl.load_workbook(filename=path2)
ws2 = wb2.create_sheet(ws1.title)

for row in ws1:
    for cell in row:
        ws2[cell.coordinate].value = cell.value

wb2.save(path2)


import pandas as pd
pd.set_option('display.height', 1000)
pd.set_option('display.max_rows', 500)
pd.set_option('display.max_columns', 500)
pd.set_option('display.width', 1000)